# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

TARGET = 235_950_000  # "235,950,00원" -> 235,950,000원으로 해석

wb = Workbook()
thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
fill_title = PatternFill("solid", fgColor="0F3E17")
fill_input = PatternFill("solid", fgColor="E1F4DF")
fill_result = PatternFill("solid", fgColor="B6CED5")
fill_header = PatternFill("solid", fgColor="CFE7D3")
fill_warn = PatternFill("solid", fgColor="FFF2CC")
font_white = Font(bold=True, color="FFFFFF", size=14)
font_bold = Font(bold=True, size=11)
won = "#,##0"
pct = "0.00%"

# ----- Sheet 1 -----
ws = wb.active
ws.title = "대가산출"

ws.merge_cells("B2:F2")
ws["B2"] = "직접인건비 + 제경비 + 기술료 역산 산출서"
ws["B2"].font = font_white
ws["B2"].fill = fill_title
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 28

ws.merge_cells("B3:F3")
ws["B3"] = "목표합계 235,950,000원 (입력 표기 235,950,00은 235,950,000으로 해석)"
ws["B3"].font = Font(size=10, italic=True, color="666666")

ws["B5"] = "【 방정식 】"
ws["B5"].font = font_bold
ws.merge_cells("B6:F10")
ws["B6"] = (
    "D = 직접인건비\n"
    "제경비 = D × α    (α = 110%~120% → 1.10~1.20)\n"
    "기술료 = (D + 제경비) × β = D(1+α)×β    (β = 20%~40% → 0.20~0.40)\n"
    "합계 S = D + Dα + D(1+α)β = D(1+α)(1+β)\n"
    "∴  D = S / ((1+α)(1+β))"
)
ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
for r in range(6, 11):
    ws.row_dimensions[r].height = 18

ws["B12"] = "【 입력 (연두칸 수정 가능) 】"
ws["B12"].font = font_bold

ws["B13"] = "항목"
ws["C13"] = "값"
ws["D13"] = "비고"
for col in ("B", "C", "D"):
    ws[f"{col}13"].fill = fill_header
    ws[f"{col}13"].font = font_bold
    ws[f"{col}13"].border = thin

ws["B14"] = "목표 합계 S (원)"
ws["C14"] = TARGET
ws["C14"].number_format = won
ws["C14"].fill = fill_input
ws["D14"] = "합계가 이 금액이 되도록 역산"

ws["B15"] = "제경비율 α"
ws["C15"] = 1.15
ws["C15"].number_format = pct
ws["C15"].fill = fill_input
ws["D15"] = "허용범위 110%~120% (기본 115%)"

ws["B16"] = "기술료율 β"
ws["C16"] = 0.30
ws["C16"].number_format = pct
ws["C16"].fill = fill_input
ws["D16"] = "허용범위 20%~40% (기본 30%)"

ws["B17"] = "α 범위 체크"
ws["C17"] = '=IF(AND(C15>=1.1,C15<=1.2),"OK","범위 이탈")'
ws["B18"] = "β 범위 체크"
ws["C18"] = '=IF(AND(C16>=0.2,C16<=0.4),"OK","범위 이탈")'

for r in range(14, 19):
    for col in ("B", "C", "D"):
        ws[f"{col}{r}"].border = thin

ws["B20"] = "【 산출 결과 (엑셀 수식) 】"
ws["B20"].font = font_bold

ws["B21"] = "항목"
ws["C21"] = "산식"
ws["D21"] = "금액 (원)"
ws["E21"] = "구성비"
for col in ("B", "C", "D", "E"):
    ws[f"{col}21"].fill = fill_header
    ws[f"{col}21"].font = font_bold
    ws[f"{col}21"].border = thin

ws["B22"] = "① 직접인건비 D"
ws["C22"] = "S / ((1+α)(1+β))"
ws["D22"] = "=C14/((1+C15)*(1+C16))"
ws["E22"] = "=D22/$D$25"

ws["B23"] = "② 제경비"
ws["C23"] = "D × α"
ws["D23"] = "=D22*C15"
ws["E23"] = "=D23/$D$25"

ws["B24"] = "③ 기술료"
ws["C24"] = "(D+제경비) × β"
ws["D24"] = "=(D22+D23)*C16"
ws["E24"] = "=D24/$D$25"

ws["B25"] = "합계 S"
ws["C25"] = "①+②+③"
ws["D25"] = "=D22+D23+D24"
ws["E25"] = "=E22+E23+E24"

for r in range(22, 26):
    ws[f"D{r}"].number_format = won
    ws[f"E{r}"].number_format = pct
    for col in ("B", "C", "D", "E"):
        ws[f"{col}{r}"].border = thin
        ws[f"{col}{r}"].fill = fill_result

ws["B25"].font = font_bold
ws["D25"].font = font_bold

ws["B27"] = "검산: 합계 − 목표"
ws["C27"] = "=D25-C14"
ws["C27"].number_format = won
ws["C27"].fill = fill_warn
ws["D27"] = '=IF(ABS(C27)<1,"일치","차이 있음")'

ws["B29"] = "【 추천 기본안 (중간요율 α=115%, β=30%) — 정수 원 단위 】"
ws["B29"].font = font_bold

# α=1.15, β=0.30 → (1+α)(1+β)=2.15×1.30=2.795
D_mid = TARGET / ((1 + 1.15) * (1 + 0.30))
oh_mid = D_mid * 1.15
D_i = int(round(D_mid))
oh_i = int(round(oh_mid))
tech_i = TARGET - D_i - oh_i

ws["B30"] = "직접인건비"
ws["C30"] = D_i
ws["B31"] = "제경비 (115%)"
ws["C31"] = oh_i
ws["B32"] = "기술료 (30%)"
ws["C32"] = tech_i
ws["B33"] = "합계"
ws["C33"] = f"=C30+C31+C32"
ws["C33"].font = font_bold

for r in range(30, 34):
    ws[f"C{r}"].number_format = won
    ws[f"B{r}"].border = thin
    ws[f"C{r}"].border = thin

ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 28
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 18

# ----- Sheet 2: scenarios -----
ws2 = wb.create_sheet("요율시나리오")
ws2.merge_cells("B2:G2")
ws2["B2"] = "α·β 조합별 직접인건비 역산표 (합계 고정 235,950,000원)"
ws2["B2"].font = font_white
ws2["B2"].fill = fill_title

headers = ["시나리오", "제경비율 α", "기술료율 β", "직접인건비 D", "제경비", "기술료", "합계"]
for i, h in enumerate(headers, start=2):
    cell = ws2.cell(3, i, h)
    cell.fill = fill_header
    cell.font = font_bold
    cell.border = thin

scenarios = [
    ("최소요율", 1.10, 0.20),
    ("제경비↓ 기술료 중간", 1.10, 0.30),
    ("제경비↓ 기술료 최대", 1.10, 0.40),
    ("중간요율 (추천)", 1.15, 0.30),
    ("제경비 중간 기술료 최대", 1.15, 0.40),
    ("제경비 최대 기술료 최소", 1.20, 0.20),
    ("제경비↑ 기술료 중간", 1.20, 0.30),
    ("최대요율", 1.20, 0.40),
]

for i, (name, a, b) in enumerate(scenarios):
    r = 4 + i
    D = TARGET / ((1 + a) * (1 + b))
    oh = D * a
    Di = int(round(D))
    ohi = int(round(oh))
    techi = TARGET - Di - ohi
    vals = [name, a, b, Di, ohi, techi, TARGET]
    for c, v in enumerate(vals, start=2):
        cell = ws2.cell(r, c, v)
        cell.border = thin
        if c in (3, 4):
            cell.number_format = pct
        elif c >= 5:
            cell.number_format = won
        if name == "중간요율 (추천)":
            cell.fill = fill_input

for col, w in zip("BCDEFG", [24, 14, 14, 16, 14, 14, 14]):
    ws2.column_dimensions[col].width = w

ws2["B13"] = "참고: 합계를 고정하면 요율(α, β)이 커질수록 직접인건비 D는 작아집니다."
ws2["B13"].font = Font(italic=True, color="666666", size=10)

# ----- Sheet 3 -----
ws3 = wb.create_sheet("추천안_상세")
ws3.merge_cells("B2:D2")
ws3["B2"] = "추천안 상세 (α=115%, β=30%)"
ws3["B2"].font = font_white
ws3["B2"].fill = fill_title

ws3["B4"] = "구분"
ws3["C4"] = "금액(원)"
ws3["D4"] = "산출근거"
for col in ("B", "C", "D"):
    ws3[f"{col}4"].fill = fill_header
    ws3[f"{col}4"].font = font_bold
    ws3[f"{col}4"].border = thin

detail_rows = [
    ("직접인건비", D_i, "S / ((1+α)(1+β))"),
    ("제경비", oh_i, "직접인건비 × 115%"),
    ("기술료", tech_i, "(직접인건비+제경비) × 30%"),
    ("합계", TARGET, "목표금액"),
]
for i, (name, amt, note) in enumerate(detail_rows):
    r = 5 + i
    ws3[f"B{r}"] = name
    ws3[f"C{r}"] = amt
    ws3[f"C{r}"].number_format = won
    ws3[f"D{r}"] = note
    for col in ("B", "C", "D"):
        ws3[f"{col}{r}"].border = thin
        if name == "합계":
            ws3[f"{col}{r}"].fill = fill_result
            ws3[f"{col}{r}"].font = font_bold

ws3["B10"] = "검산식"
ws3["B11"] = f"{D_i:,} + {oh_i:,} + {tech_i:,} = {TARGET:,}"
ws3["B12"] = (
    f"확인: D(1+α)(1+β) = {D_mid:,.4f} × 2.15 × 1.30 = {TARGET:,}"
)
ws3["B14"] = "방정식 요약"
ws3["B15"] = "S = D(1+α)(1+β) = 235,950,000"
ws3["B16"] = "D = 235,950,000 / ((1+1.15)(1+0.30)) = 235,950,000 / 2.795"

ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 40

out = r"C:\Users\user\electrical-portfolio\fee_breakdown_235950000.xlsx"
wb.save(out)
# 한글 파일명 복사
out_kr = r"C:\Users\user\electrical-portfolio\대가산출_직접인건비_제경비_기술료.xlsx"
wb.save(out_kr)
print("SAVED", out)
print("SAVED", out_kr)
print(f"추천안: D={D_i:,} / 제경비={oh_i:,} / 기술료={tech_i:,} / 합={D_i+oh_i+tech_i:,}")
for name, a, b in scenarios:
    Dd = TARGET / ((1 + a) * (1 + b))
    print(f"{name}: D={Dd:,.0f}")
